#--------------------------------------------------------------------------------------------------------------
# É muito importante para o bom funcionamento que os arquivos excel estejam na mesma pasta que o script python
# para nao temos erros ou bugs
#--------------------------------------------------------------------------------------------------------------
import openpyxl

#-----------------------------------------[ CONFIGURAÇOES ]----------------------------------------------------

planilha_inclusao = 'adere.xlsx'  # planilha que deseja modificar
planilha_inclusao_pg = 'pg'       # nome da pagina dentro da planilha

planilha_extraçao = 'adere2.xlsx' # planilha que deseja extrair informaçoes
planilha_extraçao_pg = 'pg'       # nome da pagina dentro da planilha

nova_planilha = 'adere_new.xlsx'  # Nova planilha que sera salva já com as alteraçes

#--------------------------------------------------------------------------------------------------------------

# carregando aquivo
planilha = openpyxl.load_workbook(planilha_inclusao) # Planilha de inclusao de dados
planilha2 = openpyxl.load_workbook(planilha_extraçao) # Planilha de extraçao de dados

# selecionando a pagina
pagina = planilha[planilha_inclusao_pg] # pagina da planilha de inclusao
pagina2 = planilha2[planilha_extraçao_pg] # pagina da planilha de extraçao

# Imprimindo os dados de cada linha
for linha in pagina.iter_rows(min_row=2):
    for linha2 in pagina2.iter_rows(min_row=2):
        # comparar linhas, caso for igual teremos as alteraçoes
        if linha[0].value == linha2[0].value:
            linha[4].value = linha2[1].value

print('Novo arquivo excel criado com Sucesso')

planilha.save(nova_planilha) # Salvar uma nova planilha com os dados modificados